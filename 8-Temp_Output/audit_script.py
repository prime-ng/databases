#!/usr/bin/env python3
"""
Route & Tab Inventory Audit for Prime-AI Laravel Application
Reads all route files, controller methods, and blade views.
Produces an Excel report with 3 sheets: Tab Routes, Loose Ends, All Routes.
"""

import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
APP_BASE = "/Users/bkwork/Herd/prime_ai"
OUTPUT_FILE = "/Users/bkwork/WorkFolder/1-Development/0-Git_Work/prime-ai_db/databases/8-Temp_Output/route_tab_audit_output.xlsx"

# Module → (route_prefix, route_name_prefix) as applied by RouteServiceProvider
# Derived by reading each module's RouteServiceProvider.php
MODULE_ROUTE_PREFIXES = {
    "App":               ("",                  ""),
    "Accounting":        ("accounting",         "accounting."),
    "Admission":         ("admission",          "adm."),
    "Billing":           ("billing",            "billing."),       # API module - has no web RSP prefix
    "Cafeteria":         ("cafeteria",          "cafeteria."),     # API module
    "Certificate":       ("certificate",        "certificate."),   # API module
    "Complaint":         ("complaint",          "complaint."),
    "Dashboard":         ("dashboard",          "dashboard."),     # no web prefix in RSP
    "Documentation":     ("documentation",      "documentation."),
    "EventEngine":       ("event-engine",       "event-engine."),
    "FrontOffice":       ("front-office",       "fof."),           # self-prefixed in routes file
    "GlobalMaster":      ("global-master",      "global-master."),
    "Hpc":               ("hpc",                "hpc."),           # no web prefix in RSP; self-prefixed
    "HrStaff":           ("hr-staff",           "hr-staff."),
    "Inventory":         ("inventory",          "inventory."),
    "Library":           ("library",            "library."),
    "LmsExam":           ("lms-exam",           "lms-exam."),
    "LmsHomework":       ("lms-home-work",      "lms-home-work."),
    "LmsQuests":         ("lms-quests",         "lms-quests."),
    "LmsQuiz":           ("lms-quize",          "lms-quize."),
    "Notification":      ("notification",       "notification."),
    "Payment":           ("payment",            "payment."),
    "Prime":             ("prime",              "prime."),         # API module
    "QuestionBank":      ("question-bank",      "question-bank."),
    "Recommendation":    ("recommendation",     "recommendation."),
    "Scheduler":         ("scheduler",          "scheduler."),     # API module
    "SchoolSetup":       ("school-setup",       "school-setup."),
    "SmartTimetable":    ("smart-timetable",    "smart-timetable."),  # API module, but web routes self-prefix
    "StandardTimetable": ("standard-timetable", "standard-timetable."),
    "StudentFee":        ("student-fee",        "student-fee."),
    "StudentPortal":     ("student-portal",     "student-portal."),
    "StudentProfile":    ("student-profile",    "student-profile."),
    "Syllabus":          ("syllabus",           "syllabus."),
    "SyllabusBooks":     ("syllabus-books",     "syllabus-books."),
    "SystemConfig":      ("system-config",      "system-config."),
    "TimetableFoundation": ("timetable-foundation", "timetable-foundation."),
    "Transport":         ("transport",          "transport."),
    "Vendor":            ("vendor",             "vendor."),
}

ROUTE_FILES = [
    (APP_BASE + "/routes/web.php", "App"),
    (APP_BASE + "/routes/tenant.php", "App"),
    (APP_BASE + "/Modules/Accounting/routes/web.php", "Accounting"),
    (APP_BASE + "/Modules/Admission/routes/web.php", "Admission"),
    (APP_BASE + "/Modules/Billing/routes/web.php", "Billing"),
    (APP_BASE + "/Modules/Cafeteria/routes/web.php", "Cafeteria"),
    (APP_BASE + "/Modules/Certificate/routes/web.php", "Certificate"),
    (APP_BASE + "/Modules/Complaint/routes/web.php", "Complaint"),
    (APP_BASE + "/Modules/Dashboard/routes/web.php", "Dashboard"),
    (APP_BASE + "/Modules/Documentation/routes/web.php", "Documentation"),
    (APP_BASE + "/Modules/EventEngine/routes/web.php", "EventEngine"),
    (APP_BASE + "/Modules/FrontOffice/routes/web.php", "FrontOffice"),
    (APP_BASE + "/Modules/GlobalMaster/routes/web.php", "GlobalMaster"),
    (APP_BASE + "/Modules/Hpc/routes/web.php", "Hpc"),
    (APP_BASE + "/Modules/HrStaff/routes/web.php", "HrStaff"),
    (APP_BASE + "/Modules/Inventory/routes/web.php", "Inventory"),
    (APP_BASE + "/Modules/Library/routes/web.php", "Library"),
    (APP_BASE + "/Modules/LmsExam/routes/web.php", "LmsExam"),
    (APP_BASE + "/Modules/LmsHomework/routes/web.php", "LmsHomework"),
    (APP_BASE + "/Modules/LmsQuests/routes/web.php", "LmsQuests"),
    (APP_BASE + "/Modules/LmsQuiz/routes/web.php", "LmsQuiz"),
    (APP_BASE + "/Modules/Notification/routes/web.php", "Notification"),
    (APP_BASE + "/Modules/Payment/routes/web.php", "Payment"),
    (APP_BASE + "/Modules/Prime/routes/web.php", "Prime"),
    (APP_BASE + "/Modules/QuestionBank/routes/web.php", "QuestionBank"),
    (APP_BASE + "/Modules/Recommendation/routes/web.php", "Recommendation"),
    (APP_BASE + "/Modules/Scheduler/routes/web.php", "Scheduler"),
    (APP_BASE + "/Modules/SchoolSetup/routes/web.php", "SchoolSetup"),
    (APP_BASE + "/Modules/SmartTimetable/routes/web.php", "SmartTimetable"),
    (APP_BASE + "/Modules/StandardTimetable/routes/web.php", "StandardTimetable"),
    (APP_BASE + "/Modules/StudentFee/routes/web.php", "StudentFee"),
    (APP_BASE + "/Modules/StudentPortal/routes/web.php", "StudentPortal"),
    (APP_BASE + "/Modules/StudentProfile/routes/web.php", "StudentProfile"),
    (APP_BASE + "/Modules/Syllabus/routes/web.php", "Syllabus"),
    (APP_BASE + "/Modules/SyllabusBooks/routes/web.php", "SyllabusBooks"),
    (APP_BASE + "/Modules/SystemConfig/routes/web.php", "SystemConfig"),
    (APP_BASE + "/Modules/TimetableFoundation/routes/web.php", "TimetableFoundation"),
    (APP_BASE + "/Modules/Transport/routes/web.php", "Transport"),
    (APP_BASE + "/Modules/Vendor/routes/web.php", "Vendor"),
]

# View module slug → PascalCase directory name
VIEW_MODULE_MAP = {
    "lmshomework": "LmsHomework",
    "lms-homework": "LmsHomework",
    "lms-home-work": "LmsHomework",
    "lmsexam": "LmsExam",
    "lms-exam": "LmsExam",
    "lmsquize": "LmsQuiz",
    "lms-quiz": "LmsQuiz",
    "lms-quest": "LmsQuests",
    "lms-quests": "LmsQuests",
    "school-setup": "SchoolSetup",
    "schoolsetup": "SchoolSetup",
    "system-config": "SystemConfig",
    "systemconfig": "SystemConfig",
    "smart-timetable": "SmartTimetable",
    "smarttimetable": "SmartTimetable",
    "student-profile": "StudentProfile",
    "studentprofile": "StudentProfile",
    "student-fee": "StudentFee",
    "studentfee": "StudentFee",
    "student-portal": "StudentPortal",
    "studentportal": "StudentPortal",
    "question-bank": "QuestionBank",
    "questionbank": "QuestionBank",
    "syllabus-books": "SyllabusBooks",
    "syllabusbooks": "SyllabusBooks",
    "global-master": "GlobalMaster",
    "globalmaster": "GlobalMaster",
    "timetable-foundation": "TimetableFoundation",
    "timetablefoundation": "TimetableFoundation",
    "standard-timetable": "StandardTimetable",
    "standardtimetable": "StandardTimetable",
    "hr-staff": "HrStaff",
    "hrstaff": "HrStaff",
    "front-office": "FrontOffice",
    "frontoffice": "FrontOffice",
    "event-engine": "EventEngine",
    "eventengine": "EventEngine",
    "recommendation": "Recommendation",
    "syllabus": "Syllabus",
    "transport": "Transport",
    "vendor": "Vendor",
    "library": "Library",
    "complaint": "Complaint",
    "billing": "Billing",
    "notification": "Notification",
    "dashboard": "Dashboard",
    "payment": "Payment",
    "prime": "Prime",
    "hpc": "Hpc",
    "scheduler": "Scheduler",
    "inventory": "Inventory",
    "accounting": "Accounting",
    "admission": "Admission",
    "cafeteria": "Cafeteria",
    "certificate": "Certificate",
}

# Resource method → (HTTP method, URI suffix, route suffix)
RESOURCE_MAP = {
    "index":   ("GET",    "",              ".index"),
    "create":  ("GET",    "/create",       ".create"),
    "store":   ("POST",   "",              ".store"),
    "show":    ("GET",    "/{id}",         ".show"),
    "edit":    ("GET",    "/{id}/edit",    ".edit"),
    "update":  ("PUT",    "/{id}",         ".update"),
    "destroy": ("DELETE", "/{id}",         ".destroy"),
}

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: Extract Menu Tree
# ─────────────────────────────────────────────────────────────────────────────
def extract_menu_definitions():
    """Parse MenuSyncController.php and build route_name → menu_title dict."""
    menu_file = APP_BASE + "/Modules/SystemConfig/app/Http/Controllers/MenuSyncController.php"
    menu_by_route = {}

    try:
        with open(menu_file, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
    except Exception as e:
        print(f"  WARNING: Could not read MenuSyncController: {e}")
        return menu_by_route

    # Walk through content finding 'title' => 'X' and nearby 'route' => 'Y'
    title_re = re.compile(r"'title'\s*=>\s*'([^']+)'")
    route_re = re.compile(r"'route'\s*=>\s*'([^']*)'")

    pos = 0
    while pos < len(content):
        tm = title_re.search(content, pos)
        if not tm:
            break
        title = tm.group(1)
        # Look for route within next 600 chars
        search_end = min(tm.end() + 600, len(content))
        rm = route_re.search(content, tm.end(), search_end)
        if rm:
            route_name = rm.group(1).strip()
            if route_name:
                menu_by_route[route_name] = title
        pos = tm.end()

    print(f"  Menu entries extracted: {len(menu_by_route)}")
    return menu_by_route


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: Parse Route Files
# ─────────────────────────────────────────────────────────────────────────────

def resolve_controller_file(ctrl_class, default_module):
    """
    Resolve controller class to (file_path, module_name).
    Handles full namespace or bare class name.
    """
    # Full namespace: Modules\Foo\Http\Controllers\BarController
    ns_match = re.match(r'Modules\\([A-Za-z]+)\\Http\\Controllers\\(.+)', ctrl_class)
    if ns_match:
        mod = ns_match.group(1)
        ctrl = ns_match.group(2).replace("\\", "/")
        return f"{APP_BASE}/Modules/{mod}/app/Http/Controllers/{ctrl}.php", mod

    # App\Http\Controllers\BarController
    app_match = re.match(r'App\\Http\\Controllers\\(.+)', ctrl_class)
    if app_match:
        ctrl = app_match.group(1).replace("\\", "/")
        return f"{APP_BASE}/app/Http/Controllers/{ctrl}.php", "App"

    # Bare class name — use module context
    bare = ctrl_class.strip("\\").split("\\")[-1]
    if default_module and default_module != "App":
        return f"{APP_BASE}/Modules/{default_module}/app/Http/Controllers/{bare}.php", default_module

    return f"{APP_BASE}/app/Http/Controllers/{bare}.php", "App"


def apply_module_prefix(route_name, module, local_group_prefix=""):
    """
    Apply the module's route name prefix to a local route name.
    Avoids double-prefixing when route already has the prefix.
    """
    if not route_name:
        return route_name

    mod_prefix = MODULE_ROUTE_PREFIXES.get(module, ("", ""))[1]

    # FrontOffice uses fof. prefix in its own route file — skip RSP prefix
    if module == "FrontOffice":
        # Route names in FO file already start with fof.
        return route_name

    # If route_name already starts with module prefix, don't add again
    if mod_prefix and route_name.startswith(mod_prefix):
        return route_name

    # If local group has a prefix too
    if local_group_prefix:
        full_local = local_group_prefix + route_name
        if mod_prefix:
            if full_local.startswith(mod_prefix):
                return full_local
            return mod_prefix + full_local
        return full_local

    if mod_prefix:
        return mod_prefix + route_name

    return route_name


def apply_module_uri_prefix(uri, module, local_group_prefix=""):
    """Apply module URI prefix to a route URI."""
    mod_uri_prefix = MODULE_ROUTE_PREFIXES.get(module, ("", ""))[0]

    full_uri = uri.lstrip("/")
    if local_group_prefix:
        full_uri = local_group_prefix.strip("/") + "/" + full_uri if local_group_prefix else full_uri

    # FrontOffice prefixes itself in the route file
    if module == "FrontOffice":
        return full_uri

    if mod_uri_prefix and not full_uri.startswith(mod_uri_prefix):
        return (mod_uri_prefix + "/" + full_uri).lstrip("/")

    return full_uri


def parse_route_file(filepath, module):
    """
    Parse a Laravel route file and return list of route dicts.
    Handles: Route::get/post/put/patch/delete, Route::resource,
    Route::prefix()->...->group() wrappers.
    """
    routes = []
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
    except FileNotFoundError:
        return routes, f"File not found: {filepath}"
    except Exception as e:
        return routes, str(e)

    # Remove PHP comments to avoid false matches
    content = re.sub(r'//[^\n]*', ' ', content)
    content = re.sub(r'/\*.*?\*/', ' ', content, flags=re.DOTALL)

    # ── Pattern for individual routes ───────────────────────────────────────
    # Route::METHOD('uri', [Ctrl::class, 'method'])->name('...')->...
    individual_re = re.compile(
        r"Route::(get|post|put|patch|delete)\(\s*['\"]([^'\"]+)['\"]"
        r"\s*,\s*"
        r"(?:\[?\s*([\w\\]+)::class\s*,\s*['\"](\w+)['\"]\s*\]?"  # [Ctrl::class, 'method']
        r"|['\"]([^'\"@]+)@(\w+)['\"])"  # 'Ctrl@method' style
        r"\s*\)"
        r"((?:->[\w]+\([^)]*\))*)",      # chained ->name(), ->prefix(), etc.
        re.MULTILINE
    )

    # ── Resource routes ──────────────────────────────────────────────────────
    resource_re = re.compile(
        r"Route::resource\(\s*['\"]([^'\"]+)['\"]\s*,\s*([\w\\]+)::class\s*\)"
        r"(?:->names\(\s*['\"]([^'\"]+)['\"]\s*\))?",
        re.MULTILINE
    )

    def extract_name_from_chain(chain):
        m = re.search(r"->name\(['\"]([^'\"]+)['\"]\)", chain)
        return m.group(1) if m else ""

    def add_individual_routes(text, group_uri_prefix="", group_name_prefix=""):
        """Extract individual routes from a text block with optional group context."""
        for m in individual_re.finditer(text):
            http_method = m.group(1).upper()
            uri = m.group(2).lstrip("/")
            chain = m.group(7) or ""

            if m.group(3):  # [Class::class, 'method']
                ctrl_class = m.group(3)
                ctrl_method = m.group(4)
            elif m.group(5):  # 'Class@method'
                ctrl_class = m.group(5)
                ctrl_method = m.group(6)
            else:
                continue

            # Route name from chain or derived
            local_name = extract_name_from_chain(chain)
            full_name = group_name_prefix + local_name if local_name else ""

            # Apply module prefix (unless it self-prefixes in the route file)
            if module == "FrontOffice":
                # FO self-prefixes, just combine group prefix
                final_name = full_name
                final_uri = (group_uri_prefix.strip("/") + "/" + uri).lstrip("/") if group_uri_prefix else uri
            elif module == "SmartTimetable":
                # SmartTimetable self-prefixes with smart-timetable/
                final_name = full_name
                final_uri = (group_uri_prefix.strip("/") + "/" + uri).lstrip("/") if group_uri_prefix else uri
            else:
                # Apply module prefix
                if full_name and not full_name.startswith(MODULE_ROUTE_PREFIXES.get(module, ("",""))[1]):
                    mod_name_prefix = MODULE_ROUTE_PREFIXES.get(module, ("",""))[1]
                    final_name = mod_name_prefix + full_name if mod_name_prefix else full_name
                else:
                    final_name = full_name

                mod_uri_prefix = MODULE_ROUTE_PREFIXES.get(module, ("",""))[0]
                combined_uri = (group_uri_prefix.strip("/") + "/" + uri).lstrip("/") if group_uri_prefix else uri
                if mod_uri_prefix and not combined_uri.startswith(mod_uri_prefix):
                    final_uri = (mod_uri_prefix + "/" + combined_uri).lstrip("/")
                else:
                    final_uri = combined_uri

            ctrl_file, ctrl_mod = resolve_controller_file(ctrl_class, module)
            ctrl_name = ctrl_class.split("\\")[-1]

            routes.append({
                "module": ctrl_mod,
                "http_method": http_method,
                "uri_pattern": final_uri,
                "route_name": final_name,
                "controller_class": ctrl_name,
                "controller_method": ctrl_method,
                "controller_file": ctrl_file,
            })

        # Also handle resource routes in this block
        for m in resource_re.finditer(text):
            resource_path = m.group(1).lstrip("/")
            ctrl_class = m.group(2)
            custom_name = m.group(3)

            ctrl_file, ctrl_mod = resolve_controller_file(ctrl_class, module)
            ctrl_name = ctrl_class.split("\\")[-1]

            resource_name = custom_name if custom_name else resource_path.split("/")[-1]
            full_resource_path = (group_uri_prefix.strip("/") + "/" + resource_path).lstrip("/") if group_uri_prefix else resource_path

            if module not in ("FrontOffice", "SmartTimetable"):
                mod_uri_prefix = MODULE_ROUTE_PREFIXES.get(module, ("",""))[0]
                mod_name_prefix = MODULE_ROUTE_PREFIXES.get(module, ("",""))[1]
                if mod_uri_prefix and not full_resource_path.startswith(mod_uri_prefix):
                    full_resource_path = (mod_uri_prefix + "/" + full_resource_path).lstrip("/")
            else:
                mod_name_prefix = ""

            full_resource_name = group_name_prefix + resource_name if group_name_prefix else resource_name
            if mod_name_prefix and not full_resource_name.startswith(mod_name_prefix):
                full_resource_name = mod_name_prefix + full_resource_name

            for action, (http_method, uri_suffix, name_suffix) in RESOURCE_MAP.items():
                routes.append({
                    "module": ctrl_mod,
                    "http_method": http_method,
                    "uri_pattern": full_resource_path + uri_suffix,
                    "route_name": full_resource_name + name_suffix,
                    "controller_class": ctrl_name,
                    "controller_method": action,
                    "controller_file": ctrl_file,
                })

    # ── Parse group blocks ───────────────────────────────────────────────────
    # Find Route::prefix()->...->group(function() { ... }) blocks
    # We need to handle nested groups

    def find_group_end(text, start):
        """Find the matching closing } for a ->group(function() {."""
        depth = 1
        pos = start
        while pos < len(text) and depth > 0:
            if text[pos] == '{':
                depth += 1
            elif text[pos] == '}':
                depth -= 1
            pos += 1
        return pos

    group_header_re = re.compile(
        r"Route::"
        r"(?:[\w]+\([^)]*\)->)*"  # middleware, prefix, name chains before group
        r"group\s*\(\s*function\s*\(\)\s*(?:use\s*\([^)]*\))?\s*\{",
        re.MULTILINE
    )

    # Extract groups with their prefix/name attributes
    prefix_in_chain_re = re.compile(r"->prefix\(['\"]([^'\"]+)['\"]\)")
    name_in_chain_re = re.compile(r"->name\(['\"]([^'\"]+)['\"]\)")

    # We'll process the top level and nested groups
    processed_ranges = []  # To avoid double-processing

    def process_text_for_groups(text, parent_uri_prefix="", parent_name_prefix="", depth_limit=3):
        """Recursively find group blocks and process routes."""
        if depth_limit <= 0:
            return

        pos = 0
        while pos < len(text):
            # Find next group
            gm = group_header_re.search(text, pos)
            if not gm:
                break

            # Extract full header (from last Route:: before this group)
            header_start = gm.start()
            header_end = gm.end()

            # Look back to find Route::prefix/Route::middleware chain start
            # The group call is at gm.start()
            header_text = text[max(0, header_start-300):header_end]

            # Extract prefix and name from the chain
            grp_prefixes = prefix_in_chain_re.findall(header_text)
            grp_names = name_in_chain_re.findall(header_text)

            grp_uri_prefix = grp_prefixes[-1].strip("/") if grp_prefixes else ""
            grp_name_prefix = grp_names[-1] if grp_names else ""

            # Combine with parent
            combined_uri = (parent_uri_prefix + "/" + grp_uri_prefix).strip("/") if parent_uri_prefix or grp_uri_prefix else ""
            combined_name = parent_name_prefix + grp_name_prefix

            # Find body
            body_start = header_end - 1  # starts at {
            body_end = find_group_end(text, body_start + 1)
            body = text[body_start + 1:body_end - 1]

            # Process routes in this group body
            add_individual_routes(body, combined_uri, combined_name)

            # Recurse for nested groups
            process_text_for_groups(body, combined_uri, combined_name, depth_limit - 1)

            pos = body_end

    # First pass: top-level routes (not inside any group)
    # We'll collect all group ranges and process the remainder
    group_ranges = []
    pos = 0
    temp_text = content
    while pos < len(temp_text):
        gm = group_header_re.search(temp_text, pos)
        if not gm:
            break
        header_end = gm.end()
        body_start = header_end - 1
        body_end = find_group_end(temp_text, body_start + 1)
        group_ranges.append((gm.start(), body_end))
        pos = body_end

    # Build non-group text
    non_group_parts = []
    prev = 0
    for (gs, ge) in group_ranges:
        non_group_parts.append(temp_text[prev:gs])
        prev = ge
    non_group_parts.append(temp_text[prev:])
    non_group_text = " ".join(non_group_parts)

    # Extract top-level routes
    add_individual_routes(non_group_text, "", "")

    # Extract routes inside groups
    process_text_for_groups(content, "", "", depth_limit=4)

    # Deduplicate (same http_method + uri_pattern + ctrl_method)
    seen = set()
    unique_routes = []
    for r in routes:
        key = (r["http_method"], r["uri_pattern"], r["controller_method"], r["controller_class"])
        if key not in seen:
            seen.add(key)
            unique_routes.append(r)

    return unique_routes, None


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: Controller method analysis
# ─────────────────────────────────────────────────────────────────────────────

def read_controller_method_body(ctrl_file, method_name):
    """Extract method body. Returns (body_str, error_str)."""
    if not os.path.exists(ctrl_file):
        return None, "MISSING_CONTROLLER_FILE"

    try:
        with open(ctrl_file, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
    except Exception as e:
        return None, str(e)

    method_re = re.compile(
        r"(?:public|protected|private)?\s*function\s+" + re.escape(method_name) +
        r"\s*\([^)]*\)\s*(?::\s*[\w\|\?]+)?\s*\{",
        re.MULTILINE
    )
    m = method_re.search(content)
    if not m:
        return None, "MISSING_CONTROLLER_METHOD"

    # Extract method body by matching braces
    start = m.end()
    depth = 1
    pos = start
    while pos < len(content) and depth > 0:
        if content[pos] == '{':
            depth += 1
        elif content[pos] == '}':
            depth -= 1
        pos += 1

    return content[start:pos-1], None


def analyze_get_route(ctrl_file, method_name, module):
    """
    Analyze a controller method for view(), compact(), tab indicators.
    Returns analysis dict.
    """
    result = {
        "has_view": False,
        "view_name": "",
        "view_file": "",
        "variables_passed": [],
        "is_tabbed_heuristic": False,
        "issues": [],
    }

    body, err = read_controller_method_body(ctrl_file, method_name)
    if err:
        result["issues"].append(err)
        return result

    # Check for return view(...)
    view_re = re.compile(r"return\s+view\s*\(\s*['\"]([^'\"]+)['\"]")
    view_match = view_re.search(body)

    if not view_match:
        # No view() — could be JSON, redirect, etc.
        if "return" not in body:
            result["issues"].append("EMPTY_VIEW")
        return result

    result["has_view"] = True
    view_name = view_match.group(1)
    result["view_name"] = view_name
    result["view_file"] = resolve_view_file(view_name, module)

    if result["view_file"] and not os.path.exists(result["view_file"]):
        result["issues"].append("MISSING_VIEW_FILE")

    # Extract compact() variables
    for cm in re.finditer(r"compact\(([^)]+)\)", body):
        result["variables_passed"].extend(re.findall(r"['\"](\w+)['\"]", cm.group(1)))

    # Extract ->with(['key' => ...]) variables
    for wm in re.finditer(r"->with\(\s*\[([^\]]+)\]", body):
        result["variables_passed"].extend(re.findall(r"['\"](\w+)['\"]\s*=>", wm.group(1)))

    # Deduplicate
    result["variables_passed"] = list(dict.fromkeys(result["variables_passed"]))

    # Heuristic: tabbed if method name is clearly a combined/dashboard view
    STRONGLY_TABBED_METHODS = {
        "index", "dashboard", "overview", "combined", "summary",
        "configuration", "assignment", "billing", "payment",
        "finemanagement", "fineManagement", "scholarship", "governance",
        "master", "planning", "report",
        "infrasetup", "preRequisitesSetup", "timetableConfiguration",
        "timetableMasters", "timetableRequirements", "visitorManagement",
        "communication", "registers", "compliance",
    }
    if method_name in STRONGLY_TABBED_METHODS or method_name.lower() in STRONGLY_TABBED_METHODS:
        result["is_tabbed_heuristic"] = True

    # Also heuristic-tabbed if 5+ distinct variables (real combined views)
    if len(result["variables_passed"]) >= 5:
        result["is_tabbed_heuristic"] = True

    return result


def resolve_view_file(view_name, module):
    """
    Resolve 'module::path.to.view' or 'path.to.view' to a file path.
    """
    if "::" in view_name:
        module_slug, view_path = view_name.split("::", 1)
        module_slug = module_slug.lower()
        module_dir = VIEW_MODULE_MAP.get(module_slug)
        if not module_dir:
            # Try stripping hyphens and matching
            stripped = module_slug.replace("-", "")
            for k, v in VIEW_MODULE_MAP.items():
                if k.replace("-", "") == stripped:
                    module_dir = v
                    break
        if not module_dir:
            module_dir = "".join(w.capitalize() for w in module_slug.replace("-", " ").split())
        view_parts = view_path.replace(".", "/")
        return f"{APP_BASE}/Modules/{module_dir}/resources/views/{view_parts}.blade.php"
    else:
        view_parts = view_name.replace(".", "/")
        return f"{APP_BASE}/resources/views/{view_parts}.blade.php"


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4: Extract Tab Titles from Blade Files
# ─────────────────────────────────────────────────────────────────────────────

def extract_tab_titles(blade_file):
    """
    Read blade file and extract tab info.
    Returns (tab_count, tab_titles_list, is_confirmed_tabbed).
    """
    if not blade_file or not os.path.exists(blade_file):
        return 0, [], False

    try:
        with open(blade_file, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
    except Exception:
        return 0, [], False

    # Quick check for tab indicators
    tab_indicators = [
        "nav-tabs", "tab-pane", 'data-bs-toggle="tab"', "data-bs-toggle='tab'",
        'data-toggle="tab"', "data-toggle='tab'", 'href="#', 'id="tab'
    ]
    has_tabs = any(ind in content for ind in tab_indicators)

    if not has_tabs:
        return 0, [], False

    tab_titles = []
    seen = set()

    # Pattern: nav-link attribute followed by text content
    # Handles: class="nav-link active" or class="nav-link"
    pat = re.compile(
        r'class=["\'][^"\']*nav-link[^"\']*["\'][^>]*>\s*([A-Za-z][^<\n]{1,70})',
        re.MULTILINE
    )
    for m in pat.finditer(content):
        raw = m.group(1).strip()
        # Strip any inline tags like <i>, <span>
        raw = re.sub(r'<[^>]+>', '', raw).strip()
        raw = re.sub(r'\s+', ' ', raw).strip()
        # Filter: must be readable text, not Blade directives or code
        if (raw and len(raw) >= 2 and len(raw) <= 80
                and not raw.startswith('@') and not raw.startswith('{')
                and not raw.startswith('$') and not raw.startswith('/')
                and raw not in seen):
            if re.search(r'[A-Za-z]{2}', raw):
                seen.add(raw)
                tab_titles.append(raw)

    # Fallback: look for data-bs-target="#something" nearby text
    if not tab_titles:
        pat2 = re.compile(r'nav-link[^>]*>([^<]{2,60})</', re.MULTILINE)
        for m in pat2.finditer(content):
            raw = m.group(1).strip()
            raw = re.sub(r'\s+', ' ', raw).strip()
            if (raw and len(raw) >= 2 and len(raw) <= 80
                    and not raw.startswith('@') and not raw.startswith('{')
                    and raw not in seen):
                if re.search(r'[A-Za-z]{2}', raw):
                    seen.add(raw)
                    tab_titles.append(raw)

    return len(tab_titles), tab_titles, has_tabs


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5: Write Excel
# ─────────────────────────────────────────────────────────────────────────────

def apply_header_style(ws, row_num, fill_color="1F4E79"):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def auto_size_columns(ws, max_width=65):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), max_width)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("\n=== Prime-AI Route & Tab Inventory Audit ===\n")

    # Step 1
    print("Step 1: Extracting menu definitions...")
    menu_by_route = extract_menu_definitions()

    # Step 2
    print("\nStep 2: Parsing route files...")
    all_routes = []
    modules_scanned = set()
    loose_ends = []

    for (route_file, module) in ROUTE_FILES:
        modules_scanned.add(module)
        if not os.path.exists(route_file):
            print(f"  MISSING: {route_file}")
            loose_ends.append({
                "module": module,
                "route_name": "",
                "url": route_file,
                "controller_method": "",
                "issue_type": "NO_ROUTE_FILE",
                "detail": f"Route file not found: {route_file}",
            })
            continue

        try:
            routes, err = parse_route_file(route_file, module)
        except Exception as ex:
            print(f"  EXCEPTION parsing {route_file}: {ex}")
            routes, err = [], str(ex)

        if err:
            print(f"  ERROR parsing {route_file}: {err}")
        else:
            print(f"  {module}: {len(routes)} routes found")
        all_routes.extend(routes)

    print(f"\n  Total routes collected: {len(all_routes)}")

    # Steps 3 & 4: Analyze GET routes
    print("\nStep 3+4: Analyzing GET routes for tabs and loose ends...")
    tab_routes = []
    method_cache = {}  # (ctrl_file, method) → analysis

    get_routes = [r for r in all_routes if r["http_method"] == "GET"]
    print(f"  GET routes to analyze: {len(get_routes)}")

    for route in get_routes:
        ctrl_file = route["controller_file"]
        ctrl_method = route["controller_method"]
        module = route["module"]
        cache_key = (ctrl_file, ctrl_method)

        if cache_key not in method_cache:
            try:
                analysis = analyze_get_route(ctrl_file, ctrl_method, module)
            except Exception as e:
                analysis = {
                    "has_view": False, "view_name": "", "view_file": "",
                    "variables_passed": [], "is_tabbed_heuristic": False,
                    "issues": [f"ANALYSIS_ERROR: {e}"],
                }
            method_cache[cache_key] = analysis
        else:
            analysis = method_cache[cache_key]

        # Collect issues
        route_name = route.get("route_name", "")

        if not route_name:
            loose_ends.append({
                "module": module,
                "route_name": "",
                "url": route["uri_pattern"],
                "controller_method": f"{route['controller_class']}::{ctrl_method}",
                "issue_type": "ROUTE_NO_NAME",
                "detail": f"GET {route['uri_pattern']}",
            })

        for issue in analysis["issues"]:
            if issue in ("MISSING_CONTROLLER_FILE", "MISSING_CONTROLLER_METHOD",
                         "MISSING_VIEW_FILE", "EMPTY_VIEW"):
                loose_ends.append({
                    "module": module,
                    "route_name": route_name,
                    "url": route["uri_pattern"],
                    "controller_method": f"{route['controller_class']}::{ctrl_method}",
                    "issue_type": issue,
                    "detail": ctrl_file if "CONTROLLER" in issue else analysis.get("view_file", ""),
                })

        # Check blade for tabs
        is_confirmed_tabbed = False
        tab_count = 0
        tab_titles = []

        if analysis["has_view"] and analysis["view_file"]:
            try:
                tab_count, tab_titles, blade_has_tabs = extract_tab_titles(analysis["view_file"])
                is_confirmed_tabbed = blade_has_tabs
            except Exception:
                pass

        # Include in tab routes if: confirmed by blade OR strong heuristic
        if (is_confirmed_tabbed or analysis["is_tabbed_heuristic"]) and analysis["has_view"]:
            menu_name = menu_by_route.get(route_name, "")
            view_short = analysis["view_file"].replace(APP_BASE + "/", "") if analysis["view_file"] else ""

            tab_routes.append({
                "module": module,
                "menu_name": menu_name,
                "route_name": route_name,
                "http_method": route["http_method"],
                "url_pattern": route["uri_pattern"],
                "controller_method": f"{route['controller_class']}::{ctrl_method}",
                "view_file": view_short,
                "tab_count": tab_count,
                "tab_titles": " | ".join(tab_titles) if tab_titles else "",
                "data_variables": ", ".join(analysis["variables_passed"][:20]),
                "notes": "; ".join(analysis["issues"]) if analysis["issues"] else "",
                "confirmed_tabbed": "YES" if is_confirmed_tabbed else "heuristic",
            })

    print(f"  Tabbed routes found: {len(tab_routes)}")

    # Deduplicate loose ends
    seen_issues = set()
    deduped_loose = []
    for le in loose_ends:
        key = (le["module"], le["url"], le["issue_type"], le["controller_method"])
        if key not in seen_issues:
            seen_issues.add(key)
            deduped_loose.append(le)

    print(f"  Loose ends found: {len(deduped_loose)}")

    # Step 5: Write Excel
    print(f"\nStep 5: Writing Excel...")
    wb = openpyxl.Workbook()

    # ── Sheet 1: Tab Routes ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tab Routes"

    h1 = ["Module", "Menu Name", "Route Name", "HTTP Method", "URL Pattern",
          "Controller::Method", "View File", "Tab Count", "Tab Titles (pipe-separated)",
          "Data Variables Passed", "Blade Confirmed?", "Notes"]
    ws1.append(h1)
    apply_header_style(ws1, 1)

    alt1 = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    confirmed_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    for i, r in enumerate(tab_routes):
        ws1.append([
            r["module"], r["menu_name"], r["route_name"], r["http_method"],
            r["url_pattern"], r["controller_method"], r["view_file"],
            r["tab_count"], r["tab_titles"], r["data_variables"],
            r["confirmed_tabbed"], r["notes"],
        ])
        row_fill = confirmed_fill if r["confirmed_tabbed"] == "YES" else (alt1 if i % 2 == 1 else None)
        if row_fill:
            for cell in ws1[ws1.max_row]:
                cell.fill = row_fill

    auto_size_columns(ws1)
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Loose Ends ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Loose Ends")
    h2 = ["Module", "Route Name", "URL", "Controller::Method", "Issue Type", "Detail"]
    ws2.append(h2)
    apply_header_style(ws2, 1, "C00000")

    issue_colors = {
        "MISSING_CONTROLLER_FILE":   "FFD7D7",
        "MISSING_CONTROLLER_METHOD": "FFE8CC",
        "MISSING_VIEW_FILE":         "FFF2CC",
        "EMPTY_VIEW":                "E2EFDA",
        "ROUTE_NO_NAME":             "F2F2F2",
        "NO_ROUTE_FILE":             "FFD7D7",
    }

    from collections import Counter
    issue_counter = Counter()
    for le in deduped_loose:
        issue_counter[le["issue_type"]] += 1
        ws2.append([
            le["module"], le["route_name"], le["url"],
            le["controller_method"], le["issue_type"], le["detail"],
        ])
        color = issue_colors.get(le["issue_type"], "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for cell in ws2[ws2.max_row]:
            cell.fill = fill

    auto_size_columns(ws2)
    ws2.freeze_panes = "A2"

    # ── Sheet 3: All Routes ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("All Routes")
    h3 = ["Module", "Route Name", "HTTP Method", "URL", "Controller::Method", "Is Tabbed?"]
    ws3.append(h3)
    apply_header_style(ws3, 1, "375623")

    tabbed_keys = set(
        (r["url_pattern"], r["controller_method"])
        for r in tab_routes
    )

    alt3 = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    tab3 = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for i, r in enumerate(all_routes):
        cm = f"{r['controller_class']}::{r['controller_method']}"
        is_tab = "YES" if (r["uri_pattern"], cm) in tabbed_keys else ""
        ws3.append([r["module"], r.get("route_name",""), r["http_method"], r["uri_pattern"], cm, is_tab])
        if is_tab:
            for cell in ws3[ws3.max_row]:
                cell.fill = tab3
        elif i % 2 == 1:
            for cell in ws3[ws3.max_row]:
                cell.fill = alt3

    auto_size_columns(ws3)
    ws3.freeze_panes = "A2"

    # ── Save ────────────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)

    # ── Summary ─────────────────────────────────────────────────────────────
    confirmed_count = sum(1 for r in tab_routes if r["confirmed_tabbed"] == "YES")
    print("\n" + "="*55)
    print(f"Modules scanned:           {len(modules_scanned)}")
    print(f"Routes found:              {len(all_routes)}")
    print(f"Tabbed routes found:       {len(tab_routes)}")
    print(f"  - Blade confirmed tabs:  {confirmed_count}")
    print(f"  - Heuristic only:        {len(tab_routes) - confirmed_count}")
    print(f"Loose ends found:          {len(deduped_loose)}")
    for k, v in sorted(issue_counter.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")
    print(f"Output written to:         {OUTPUT_FILE}")
    print("="*55 + "\n")


if __name__ == "__main__":
    main()
